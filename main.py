"""
Paie Burkina Faso - Application Android (KivyMD)
Génère un APK via Buildozer
"""

import json
import os
from math import floor
from datetime import datetime

from kivy.config import Config
Config.set('graphics', 'width', '360')
Config.set('graphics', 'height', '640')

from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.recycleview import RecycleView
from kivy.uix.recycleboxlayout import RecycleBoxLayout
from kivy.uix.recycleview.views import RecycleDataViewBehavior
from kivy.uix.behaviors import FocusBehavior
from kivy.uix.popup import Popup
from kivy.properties import StringProperty, NumericProperty, ListProperty, ObjectProperty, BooleanProperty
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.core.window import Window

from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.button import MDRaisedButton, MDIconButton, MDFlatButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivymd.uix.card import MDCard
from kivymd.uix.dialog import MDDialog
from kivymd.uix.snackbar import MDSnackbar
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.list import MDList, OneLineListItem, TwoLineListItem
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.selectioncontrol import MDCheckbox

# ---------------------------------------------------------------------------
# PARAMÈTRES PAR DÉFAUT
# ---------------------------------------------------------------------------
DEFAULT_SETTINGS = {
    "mois": "DECEMBER",
    "annee": 2026,
    "cnss_employe_pct": 0.055,
    "plafond_cnss": 800000,
    "max_cnss_employe": 44000,
    "plafond_fiscal_pct": 0.08,
    "abattement_cadre": 0.20,
    "abattement_non_cadre": 0.25,
    "plafond_exon_logement": 75000,
    "plafond_exon_fonction": 50000,
    "seuil_exon_fonction": 30000,
    "plafond_exon_transport": 30000,
    "seuil_exon_transport": 30000,
    "tpa_pct": 0.03,
    "cnss_patronale_pct": 0.16,
    "retenue_obligatoire_pct": 0.01,
    "reduction_0": 1.00,
    "reduction_1": 0.92,
    "reduction_2": 0.90,
    "reduction_3": 0.88,
    "reduction_4": 0.86,
    # Tranches IUTS
    "tranches": [
        {"plafond": 10000, "taux": 0.0, "cumul": 0},
        {"plafond": 20000, "taux": 0.0, "cumul": 0},
        {"plafond": 30000, "taux": 0.0, "cumul": 0},
        {"plafond": 50000, "taux": 0.121, "cumul": 0},
        {"plafond": 80000, "taux": 0.139, "cumul": 2420},
        {"plafond": 120000, "taux": 0.157, "cumul": 6590},
        {"plafond": 170000, "taux": 0.184, "cumul": 12870},
        {"plafond": 250000, "taux": 0.217, "cumul": 22070},
        {"plafond": 999999999, "taux": 0.25, "cumul": 39430},
    ]
}

# ---------------------------------------------------------------------------
# DONNÉES EMPLOYÉS PAR DÉFAUT
# ---------------------------------------------------------------------------
DEFAULT_EMPLOYEES = [
    {"no": 1, "nom": "KAF", "prenom": "JUILLE", "classification": "AUTRE",
     "sal_base": 101933, "prim_anc": 0, "heure_sup": 0, "sursalaire": 0,
     "gratif": 0, "indem": 0, "caisse": 30000, "logement": 15000,
     "fonction": 20000, "transport": 0, "charges": 2, "retenue_pret": 64000},
    {"no": 2, "nom": "DOLOP", "prenom": "DILL", "classification": "AUTRE",
     "sal_base": 500000, "prim_anc": 5, "heure_sup": 5, "sursalaire": 5,
     "gratif": 5, "indem": 10000, "caisse": 56000, "logement": 25000,
     "fonction": 23000, "transport": 0, "charges": 5, "retenue_pret": 20000},
    {"no": 3, "nom": "DERME", "prenom": "DERM", "classification": "AUTRE",
     "sal_base": 500000, "prim_anc": 0, "heure_sup": 0, "sursalaire": 0,
     "gratif": 0, "indem": 0, "caisse": 62000, "logement": 62000,
     "fonction": 65000, "transport": 0, "charges": 1, "retenue_pret": 0},
    {"no": 4, "nom": "DERMO", "prenom": "RTE", "classification": "AUTRE",
     "sal_base": 19999, "prim_anc": 0, "heure_sup": 0, "sursalaire": 0,
     "gratif": 0, "indem": 0, "caisse": 0, "logement": 10000,
     "fonction": 20000, "transport": 0, "charges": 1, "retenue_pret": 0},
    {"no": 5, "nom": "LOP", "prenom": "JOP", "classification": "AUTRE",
     "sal_base": 20000, "prim_anc": 0, "heure_sup": 0, "sursalaire": 0,
     "gratif": 0, "indem": 0, "caisse": 10000, "logement": 10000,
     "fonction": 10000, "transport": 0, "charges": 2, "retenue_pret": 0},
]

# ---------------------------------------------------------------------------
# FONCTIONS DE CALCUL
# ---------------------------------------------------------------------------
def round_fcfa(val):
    return round(val)

def compute_employee(emp, settings):
    """Calcule toute la paie d'un employé."""
    s = settings

    # --- Éléments saisis ---
    sal_base = float(emp.get("sal_base", 0) or 0)
    prim_anc = float(emp.get("prim_anc", 0) or 0)
    heure_sup = float(emp.get("heure_sup", 0) or 0)
    sursalaire = float(emp.get("sursalaire", 0) or 0)
    gratif = float(emp.get("gratif", 0) or 0)
    indem = float(emp.get("indem", 0) or 0)
    caisse = float(emp.get("caisse", 0) or 0)
    logement = float(emp.get("logement", 0) or 0)
    fonction = float(emp.get("fonction", 0) or 0)
    transport = float(emp.get("transport", 0) or 0)
    charges = int(emp.get("charges", 0) or 0)
    retenue_pret = float(emp.get("retenue_pret", 0) or 0)
    classification = emp.get("classification", "AUTRE")

    # --- Rémunération Totale ---
    # Fidélité stricte : SUM(F:O) = Sal Base + Prim + Heure Sup + Sursalaire + Gratif + Indem + Caisse + Log + Fct° + Trpt
    rem_total = sal_base + prim_anc + heure_sup + sursalaire + gratif + indem + caisse + logement + fonction + transport

    # --- CNSS Employé ---
    if rem_total <= s["plafond_cnss"]:
        cnss_employe = round_fcfa(rem_total * s["cnss_employe_pct"])
    else:
        cnss_employe = s["max_cnss_employe"]

    # --- Plafond fiscal ---
    plafond_fiscal = s["plafond_fiscal_pct"] * (sal_base + prim_anc + heure_sup + sursalaire)

    # --- Salaire Brut ---
    if cnss_employe >= plafond_fiscal:
        sal_brut = round_fcfa(rem_total - plafond_fiscal)
    else:
        sal_brut = round_fcfa(rem_total - cnss_employe)

    # --- Abattement ---
    base_abatt = sal_base + prim_anc + heure_sup + sursalaire
    if classification == "CADRE":
        abattement = round_fcfa(s["abattement_cadre"] * base_abatt)
    else:
        abattement = round_fcfa(s["abattement_non_cadre"] * base_abatt)

    # --- Exonérations ---
    # Exonération "Logement" (colonne W) — FIDÉLITÉ STRICTE AU FICHIER EXCEL:
    # Dans l'original, la formule W pointe vers M (Caisse) et non N (Logement)
    vingt_pct_brut = 0.20 * sal_brut
    if vingt_pct_brut <= caisse:  # <= M (Caisse) dans l'original
        if vingt_pct_brut <= s["plafond_exon_logement"]:
            exon_log = vingt_pct_brut
        else:
            exon_log = s["plafond_exon_logement"]
    else:
        if caisse >= s["plafond_exon_logement"]:  # M >= 75000 dans l'original
            exon_log = s["plafond_exon_logement"]
        else:
            exon_log = caisse

    # Exonération "Fonction" (colonne X) — FIDÉLITÉ STRICTE AU FICHIER EXCEL:
    # Dans l'original, X pointe vers N (Logement) et non O (Fonction)
    cinq_pct_brut = 0.05 * sal_brut
    if cinq_pct_brut <= logement:  # <= N (Logement) dans l'original
        if cinq_pct_brut <= s["plafond_exon_fonction"]:
            exon_fct = cinq_pct_brut
        else:
            exon_fct = s["plafond_exon_fonction"]
    else:
        if logement >= s["seuil_exon_fonction"]:  # N >= 30000 dans l'original
            exon_fct = s["plafond_exon_fonction"]
        else:
            exon_fct = logement

    # Exonération "Transport" (colonne Y) — FIDÉLITÉ STRICTE AU FICHIER EXCEL:
    # Dans l'original, Y pointe vers O (Fonction) et non P (Transport)
    if cinq_pct_brut <= fonction:  # <= O (Fonction) dans l'original
        if cinq_pct_brut <= s["plafond_exon_transport"]:
            exon_trpt = cinq_pct_brut
        else:
            exon_trpt = s["plafond_exon_transport"]
    else:
        if fonction >= s["seuil_exon_transport"]:  # O >= 30000 dans l'original
            exon_trpt = s["plafond_exon_transport"]
        else:
            exon_trpt = fonction

    total_exon = abattement + exon_log + exon_fct + exon_trpt

    # --- Base imposable ---
    base_imp = floor((sal_brut - total_exon) / 100) * 100  # ROUNDDOWN(..., -2)

    # --- IUTS Brut (progressif) ---
    tranches = s["tranches"]
    iuts_brut = 0
    if base_imp < tranches[0]["plafond"]:
        iuts_brut = 0
    else:
        for i in range(len(tranches)):
            if base_imp < tranches[i]["plafond"]:
                prev = tranches[i-1]
                iuts_brut = (base_imp - prev["plafond"]) * prev["taux"] + prev["cumul"]
                break
        else:
            # Au-delà de la dernière tranche
            last = tranches[-2]
            iuts_brut = (base_imp - last["plafond"]) * last["taux"] + last["cumul"]
            # Tranche 9
            last2 = tranches[-1]
            iuts_brut = (base_imp - tranches[-2]["plafond"]) * tranches[-2]["taux"] + tranches[-2]["cumul"]
            if base_imp > tranches[-2]["plafond"]:
                iuts_brut = (base_imp - tranches[-2]["plafond"]) * 0.25 + 39430

    # Correction formule exacte de l'original
    if base_imp < 10000:
        iuts_brut = 0
    elif base_imp < 20000:
        iuts_brut = 0
    elif base_imp < 30000:
        iuts_brut = 0
    elif base_imp < 50000:
        iuts_brut = (base_imp - 30000) * 0.121
    elif base_imp < 80000:
        iuts_brut = (base_imp - 50000) * 0.139 + 2420
    elif base_imp < 120000:
        iuts_brut = (base_imp - 80000) * 0.157 + 6590
    elif base_imp < 170000:
        iuts_brut = (base_imp - 120000) * 0.184 + 12870
    elif base_imp <= 250000:
        iuts_brut = (base_imp - 170000) * 0.217 + 22070
    else:
        iuts_brut = (base_imp - 250000) * 0.25 + 39430

    # --- IUTS Net (réduction selon charges) ---
    if charges == 0:
        reduc = s["reduction_0"]
    elif charges == 1:
        reduc = s["reduction_1"]
    elif charges == 2:
        reduc = s["reduction_2"]
    elif charges == 3:
        reduc = s["reduction_3"]
    else:
        reduc = s["reduction_4"]

    iuts_net = round_fcfa(iuts_brut * reduc)

    # --- Salaire Net ---
    sal_net = rem_total - cnss_employe - iuts_net

    # --- Retenue 1% ---
    retenue_1pct = round_fcfa(sal_net * s["retenue_obligatoire_pct"])

    # --- Net Perçu ---
    net_percu = sal_net - retenue_1pct - retenue_pret

    # --- Charges patronales ---
    tpa = round_fcfa(rem_total * s["tpa_pct"])
    cnss_patronale = round_fcfa(rem_total * s["cnss_patronale_pct"])
    total_charges_patronales = tpa + cnss_patronale

    # --- Coûts ---
    couts = rem_total + total_charges_patronales

    # --- Totaux ---
    cnss_total = cnss_patronale + cnss_employe
    iuts_tpa = tpa + iuts_net

    return {
        "no": emp["no"],
        "nom": emp["nom"],
        "prenom": emp["prenom"],
        "classification": classification,
        "sal_base": sal_base, "prim_anc": prim_anc, "heure_sup": heure_sup,
        "sursalaire": sursalaire, "gratif": gratif, "indem": indem,
        "caisse": caisse, "logement": logement, "fonction": fonction,
        "transport": transport, "charges": charges, "retenue_pret": retenue_pret,
        "rem_total": rem_total,
        "cnss_employe": cnss_employe,
        "plafond_fiscal": plafond_fiscal,
        "sal_brut": sal_brut,
        "abattement": abattement,
        "exon_log": exon_log, "exon_fct": exon_fct, "exon_trpt": exon_trpt,
        "total_exon": total_exon,
        "base_imp": base_imp,
        "iuts_brut": iuts_brut,
        "iuts_net": iuts_net,
        "sal_net": sal_net,
        "retenue_1pct": retenue_1pct,
        "net_percu": net_percu,
        "tpa": tpa,
        "cnss_patronale": cnss_patronale,
        "total_charges_patronales": total_charges_patronales,
        "couts": couts,
        "cnss_total": cnss_total,
        "iuts_tpa": iuts_tpa,
    }

def compute_totals(results):
    """Calcule les totaux de toutes les colonnes."""
    keys = [
        "sal_base", "prim_anc", "heure_sup", "sursalaire", "gratif", "indem",
        "caisse", "logement", "fonction", "transport", "rem_total",
        "cnss_employe", "plafond_fiscal", "sal_brut", "abattement",
        "exon_log", "exon_fct", "exon_trpt", "total_exon", "base_imp",
        "iuts_brut", "iuts_net", "sal_net", "retenue_1pct", "retenue_pret",
        "net_percu", "tpa", "cnss_patronale", "total_charges_patronales",
        "couts", "cnss_total", "iuts_tpa"
    ]
    totals = {k: sum(r[k] for r in results) for k in keys}
    return totals

def generate_journal_entries(totals):
    """Génère les écritures comptables SYSCOHADA."""
    entries = [
        {"debit": "661100", "credit": None, "libelle": "SAL DE BASE", "montant": totals["sal_base"]},
        {"debit": "661200", "credit": None, "libelle": "PRIMES ANCIEN ET GRATIFICATIONS", "montant": totals["prim_anc"] + totals["gratif"]},
        {"debit": "661800", "credit": None, "libelle": "HEURES SUPL ET SURSALAIRE", "montant": totals["heure_sup"] + totals["sursalaire"]},
        {"debit": "663800", "credit": None, "libelle": "INDEMNITE DE CAISSE", "montant": totals["caisse"]},
        {"debit": "663100", "credit": None, "libelle": "INDEMNITE LOGEMENT", "montant": totals["logement"]},
        {"debit": "663200", "credit": None, "libelle": "INDEMNITE DE FONCTION", "montant": totals["fonction"]},
        {"debit": "663400", "credit": None, "libelle": "INDEMNITE DE TRANSPORT", "montant": totals["transport"]},
        {"debit": None, "credit": "422000", "libelle": "SALAIRE NET VIREMENT", "montant": totals["net_percu"]},
        {"debit": None, "credit": "447220", "libelle": "ROSALAIRE", "montant": totals["retenue_1pct"]},
        {"debit": None, "credit": "421000", "libelle": "RETENUE AVANCE /SALAIRES", "montant": totals["retenue_pret"]},
        {"debit": None, "credit": "431300", "libelle": "CNSS EMPLOYE", "montant": totals["cnss_employe"]},
        {"debit": None, "credit": "447210", "libelle": "IUTS", "montant": totals["iuts_net"]},
    ]

    sub1_debit = sum(e["montant"] for e in entries if e["debit"])
    sub1_credit = sum(e["montant"] for e in entries if e["credit"])

    patronal = [
        {"debit": "664100", "credit": None, "libelle": "CNSS PATRONALE", "montant": totals["cnss_patronale"]},
        {"debit": None, "credit": "431300", "libelle": "CNSS PATRONALE", "montant": totals["cnss_patronale"]},
        {"debit": "664200", "credit": None, "libelle": "TPA", "montant": totals["tpa"]},
        {"debit": None, "credit": "447230", "libelle": "TPA", "montant": totals["tpa"]},
    ]

    sub2_debit = sum(e["montant"] for e in patronal if e["debit"])
    sub2_credit = sum(e["montant"] for e in patronal if e["credit"])

    return {
        "entries": entries,
        "sub1_debit": sub1_debit,
        "sub1_credit": sub1_credit,
        "patronal": patronal,
        "sub2_debit": sub2_debit,
        "sub2_credit": sub2_credit,
        "grand_total": sub1_debit + sub2_debit
    }

# ---------------------------------------------------------------------------
# SCREENS
# ---------------------------------------------------------------------------

class SettingsScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        toolbar = MDTopAppBar(title="Paramètres", elevation=4)
        toolbar.left_action_items = [["arrow-left", lambda x: self.manager.switch_screen("menu")]]
        layout.add_widget(toolbar)

        scroll = MDScrollView()
        grid = MDGridLayout(cols=1, spacing=dp(10), padding=dp(16), size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))

        self.fields = {}
        app = MDApp.get_running_app()
        s = app.settings

        params = [
            ("mois", "Mois", s["mois"]),
            ("annee", "Année", str(s["annee"])),
            ("cnss_employe_pct", "CNSS Employé (%)", str(s["cnss_employe_pct"]*100)),
            ("plafond_cnss", "Plafond CNSS (FCFA)", str(s["plafond_cnss"])),
            ("max_cnss_employe", "Max CNSS Employé (FCFA)", str(s["max_cnss_employe"])),
            ("plafond_fiscal_pct", "Plafond Fiscal (%)", str(s["plafond_fiscal_pct"]*100)),
            ("abattement_cadre", "Abattement Cadre (%)", str(s["abattement_cadre"]*100)),
            ("abattement_non_cadre", "Abattement Non-Cadre (%)", str(s["abattement_non_cadre"]*100)),
            ("plafond_exon_logement", "Plafond Exon Logement (FCFA)", str(s["plafond_exon_logement"])),
            ("plafond_exon_fonction", "Plafond Exon Fonction (FCFA)", str(s["plafond_exon_fonction"])),
            ("seuil_exon_fonction", "Seuil Exon Fonction (FCFA)", str(s["seuil_exon_fonction"])),
            ("plafond_exon_transport", "Plafond Exon Transport (FCFA)", str(s["plafond_exon_transport"])),
            ("seuil_exon_transport", "Seuil Exon Transport (FCFA)", str(s["seuil_exon_transport"])),
            ("tpa_pct", "TPA (%)", str(s["tpa_pct"]*100)),
            ("cnss_patronale_pct", "CNSS Patronale (%)", str(s["cnss_patronale_pct"]*100)),
            ("retenue_obligatoire_pct", "Retenue Obligatoire (%)", str(s["retenue_obligatoire_pct"]*100)),
        ]

        for key, label, value in params:
            tf = MDTextField(hint_text=label, text=value, mode="rectangle")
            self.fields[key] = tf
            grid.add_widget(tf)

        scroll.add_widget(grid)
        layout.add_widget(scroll)

        btn = MDRaisedButton(text="Enregistrer", pos_hint={"center_x": 0.5}, size_hint=(0.8, None))
        btn.bind(on_release=self.save_settings)
        layout.add_widget(btn)

        self.add_widget(layout)

    def save_settings(self, *args):
        app = MDApp.get_running_app()
        try:
            app.settings["mois"] = self.fields["mois"].text
            app.settings["annee"] = int(self.fields["annee"].text)
            app.settings["cnss_employe_pct"] = float(self.fields["cnss_employe_pct"].text) / 100
            app.settings["plafond_cnss"] = float(self.fields["plafond_cnss"].text)
            app.settings["max_cnss_employe"] = float(self.fields["max_cnss_employe"].text)
            app.settings["plafond_fiscal_pct"] = float(self.fields["plafond_fiscal_pct"].text) / 100
            app.settings["abattement_cadre"] = float(self.fields["abattement_cadre"].text) / 100
            app.settings["abattement_non_cadre"] = float(self.fields["abattement_non_cadre"].text) / 100
            app.settings["plafond_exon_logement"] = float(self.fields["plafond_exon_logement"].text)
            app.settings["plafond_exon_fonction"] = float(self.fields["plafond_exon_fonction"].text)
            app.settings["seuil_exon_fonction"] = float(self.fields["seuil_exon_fonction"].text)
            app.settings["plafond_exon_transport"] = float(self.fields["plafond_exon_transport"].text)
            app.settings["seuil_exon_transport"] = float(self.fields["seuil_exon_transport"].text)
            app.settings["tpa_pct"] = float(self.fields["tpa_pct"].text) / 100
            app.settings["cnss_patronale_pct"] = float(self.fields["cnss_patronale_pct"].text) / 100
            app.settings["retenue_obligatoire_pct"] = float(self.fields["retenue_obligatoire_pct"].text) / 100
            app.save_data()
            MDSnackbar(text="Paramètres enregistrés !").open()
        except Exception as e:
            MDSnackbar(text=f"Erreur: {str(e)}").open()


class EmployeeEditDialog(MDBoxLayout):
    def __init__(self, employee=None, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.spacing = dp(10)
        self.size_hint_y = None
        self.height = dp(600)

        self.emp = employee or {}
        self.fields = {}

        fields_def = [
            ("nom", "Nom"),
            ("prenom", "Prénom"),
            ("classification", "Classification (CADRE/AUTRE)"),
            ("sal_base", "Sal de Base"),
            ("prim_anc", "Prime Ancienneté"),
            ("heure_sup", "Heure Sup"),
            ("sursalaire", "Sursalaire"),
            ("gratif", "Gratification"),
            ("indem", "Indemnité"),
            ("caisse", "Caisse"),
            ("logement", "Logement"),
            ("fonction", "Fonction"),
            ("transport", "Transport"),
            ("charges", "Nombre de Charges"),
            ("retenue_pret", "Retenue Prêt/Avances"),
        ]

        for key, hint in fields_def:
            val = str(self.emp.get(key, ""))
            if key in ["sal_base", "prim_anc", "heure_sup", "sursalaire", "gratif", "indem", "caisse", "logement", "fonction", "transport", "retenue_pret"]:
                val = str(int(self.emp.get(key, 0)))
            tf = MDTextField(hint_text=hint, text=val, mode="rectangle")
            self.fields[key] = tf
            self.add_widget(tf)


class DataEntryScreen(MDScreen):
    dialog = None

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        toolbar = MDTopAppBar(title="Saisie Employés", elevation=4)
        toolbar.left_action_items = [["arrow-left", lambda x: self.manager.switch_screen("menu")]]
        toolbar.right_action_items = [["plus", lambda x: self.open_add_dialog()]]
        layout.add_widget(toolbar)

        self.list_view = MDList()
        scroll = MDScrollView()
        scroll.add_widget(self.list_view)
        layout.add_widget(scroll)

        self.add_widget(layout)
        Clock.schedule_once(self.refresh_list, 0.5)

    def refresh_list(self, *args):
        self.list_view.clear_widgets()
        app = MDApp.get_running_app()
        for emp in app.employees:
            item = TwoLineListItem(
                text=f"{emp['no']}. {emp['nom']} {emp['prenom']} ({emp['classification']})",
                secondary_text=f"Base: {int(emp['sal_base']):,} FCFA | Charges: {emp['charges']}",
                on_release=lambda x, e=emp: self.open_edit_dialog(e)
            )
            self.list_view.add_widget(item)

    def open_add_dialog(self):
        self.show_dialog(None)

    def open_edit_dialog(self, emp):
        self.show_dialog(emp)

    def show_dialog(self, emp):
        content = EmployeeEditDialog(employee=emp)
        self.dialog = MDDialog(
            title="Employé" if emp is None else f"Modifier {emp['nom']}",
            type="custom",
            content_cls=content,
            buttons=[
                MDFlatButton(text="ANNULER", on_release=lambda x: self.dialog.dismiss()),
                MDRaisedButton(text="ENREGISTRER", on_release=lambda x: self.save_employee(content, emp)),
            ],
        )
        self.dialog.open()

    def save_employee(self, content, existing):
        app = MDApp.get_running_app()
        try:
            data = {k: v.text for k, v in content.fields.items()}
            emp = {
                "no": existing["no"] if existing else len(app.employees) + 1,
                "nom": data["nom"].upper(),
                "prenom": data["prenom"].upper(),
                "classification": data["classification"].upper(),
                "sal_base": float(data["sal_base"] or 0),
                "prim_anc": float(data["prim_anc"] or 0),
                "heure_sup": float(data["heure_sup"] or 0),
                "sursalaire": float(data["sursalaire"] or 0),
                "gratif": float(data["gratif"] or 0),
                "indem": float(data["indem"] or 0),
                "caisse": float(data["caisse"] or 0),
                "logement": float(data["logement"] or 0),
                "fonction": float(data["fonction"] or 0),
                "transport": float(data["transport"] or 0),
                "charges": int(data["charges"] or 0),
                "retenue_pret": float(data["retenue_pret"] or 0),
            }
            if existing:
                idx = app.employees.index(existing)
                app.employees[idx] = emp
            else:
                app.employees.append(emp)
            app.save_data()
            self.dialog.dismiss()
            self.refresh_list()
            MDSnackbar(text="Employé enregistré !").open()
        except Exception as e:
            MDSnackbar(text=f"Erreur: {str(e)}").open()


class PayrollScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        toolbar = MDTopAppBar(title="Payroll Data", elevation=4)
        toolbar.left_action_items = [["arrow-left", lambda x: self.manager.switch_screen("menu")]]
        layout.add_widget(toolbar)

        scroll = MDScrollView()
        self.grid = MDGridLayout(cols=1, spacing=dp(4), padding=dp(8), size_hint_y=None)
        self.grid.bind(minimum_height=self.grid.setter("height"))
        scroll.add_widget(self.grid)
        layout.add_widget(scroll)

        self.add_widget(layout)
        self.bind(on_enter=self.refresh)

    def refresh(self, *args):
        self.grid.clear_widgets()
        app = MDApp.get_running_app()
        results = [compute_employee(emp, app.settings) for emp in app.employees]
        totals = compute_totals(results)

        # En-tête
        self.add_card("PÉRIODE", f"{app.settings['mois']} {app.settings['annee']}", bg=(0.12, 0.24, 0.47, 1))

        for r in results:
            card = MDCard(orientation="vertical", size_hint_y=None, height=dp(420), padding=dp(12), elevation=2)
            card.add_widget(MDLabel(text=f"{r['no']}. {r['nom']} {r['prenom']}", theme_text_color="Primary", font_style="H6"))

            info = [
                ("Rém. Total", r["rem_total"]),
                ("CNSS Employé", r["cnss_employe"]),
                ("Plafond Fiscal", r["plafond_fiscal"]),
                ("Sal. Brut", r["sal_brut"]),
                ("Abattement", r["abattement"]),
                ("Exonérations", r["total_exon"]),
                ("Base Imposable", r["base_imp"]),
                ("IUTS Net", r["iuts_net"]),
                ("Salaire Net", r["sal_net"]),
                ("Retenue 1%", r["retenue_1pct"]),
                ("Retenue Prêt", r["retenue_pret"]),
                ("NET PERÇU", r["net_percu"]),
                ("TPA 3%", r["tpa"]),
                ("CNSS Patronale", r["cnss_patronale"]),
                ("COÛTS", r["couts"]),
            ]

            for label, val in info:
                bold = True if label == "NET PERÇU" or label == "COÛTS" else False
                color = (0.75, 0, 0, 1) if bold else (0, 0, 0, 0.7)
                lbl = MDLabel(text=f"{label}: {val:,.0f} FCFA", theme_text_color="Custom", text_color=color, bold=bold)
                card.add_widget(lbl)

            self.grid.add_widget(card)

        # TOTAL
        self.add_card("TOTAL RÉM. TOTAL", totals["rem_total"], bg=(0.27, 0.54, 0.27, 1))
        self.add_card("TOTAL NET PERÇU", totals["net_percu"], bg=(0.27, 0.54, 0.27, 1))
        self.add_card("TOTAL COÛTS", totals["couts"], bg=(0.75, 0, 0, 1))

    def add_card(self, label, value, bg=(1,1,1,1)):
        card = MDCard(orientation="horizontal", size_hint_y=None, height=dp(50), padding=dp(12), elevation=1)
        card.md_bg_color = bg
        lbl = MDLabel(text=f"{label}: {value:,.0f} FCFA", theme_text_color="Custom", text_color=(1,1,1,1), bold=True, halign="center")
        card.add_widget(lbl)
        self.grid.add_widget(card)


class JournalScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        toolbar = MDTopAppBar(title="Journal Entries", elevation=4)
        toolbar.left_action_items = [["arrow-left", lambda x: self.manager.switch_screen("menu")]]
        layout.add_widget(toolbar)

        scroll = MDScrollView()
        self.grid = MDGridLayout(cols=1, spacing=dp(4), padding=dp(8), size_hint_y=None)
        self.grid.bind(minimum_height=self.grid.setter("height"))
        scroll.add_widget(self.grid)
        layout.add_widget(scroll)

        self.add_widget(layout)
        self.bind(on_enter=self.refresh)

    def refresh(self, *args):
        self.grid.clear_widgets()
        app = MDApp.get_running_app()
        results = [compute_employee(emp, app.settings) for emp in app.employees]
        totals = compute_totals(results)
        journal = generate_journal_entries(totals)

        # Écritures
        for e in journal["entries"]:
            if e["debit"]:
                text = f"{e['debit']} | {e['libelle']} | DÉBIT: {e['montant']:,.0f}"
                color = (0, 0.4, 0, 1)
            else:
                text = f"{e['credit']} | {e['libelle']} | CRÉDIT: {e['montant']:,.0f}"
                color = (0.7, 0, 0, 1)
            card = MDCard(size_hint_y=None, height=dp(40), padding=dp(8), elevation=1)
            card.add_widget(MDLabel(text=text, theme_text_color="Custom", text_color=color, font_style="Body2"))
            self.grid.add_widget(card)

        self.add_total_card(f"SOUS-TOTAL 1 | DÉBIT: {journal['sub1_debit']:,.0f} | CRÉDIT: {journal['sub1_credit']:,.0f}", (0.12, 0.24, 0.47, 1))

        for e in journal["patronal"]:
            if e["debit"]:
                text = f"{e['debit']} | {e['libelle']} | DÉBIT: {e['montant']:,.0f}"
            else:
                text = f"{e['credit']} | {e['libelle']} | CRÉDIT: {e['montant']:,.0f}"
            card = MDCard(size_hint_y=None, height=dp(40), padding=dp(8), elevation=1)
            card.add_widget(MDLabel(text=text, theme_text_color="Custom", text_color=(0,0,0,0.7), font_style="Body2"))
            self.grid.add_widget(card)

        self.add_total_card(f"SOUS-TOTAL 2 | DÉBIT: {journal['sub2_debit']:,.0f} | CRÉDIT: {journal['sub2_credit']:,.0f}", (0.12, 0.24, 0.47, 1))
        self.add_total_card(f"GRAND TOTAL: {journal['grand_total']:,.0f} FCFA", (0.75, 0, 0, 1))

    def add_total_card(self, text, bg):
        card = MDCard(size_hint_y=None, height=dp(50), padding=dp(8), elevation=2)
        card.md_bg_color = bg
        card.add_widget(MDLabel(text=text, theme_text_color="Custom", text_color=(1,1,1,1), bold=True, halign="center"))
        self.grid.add_widget(card)


class MenuScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = MDBoxLayout(orientation="vertical")
        toolbar = MDTopAppBar(title="Paie Burkina Faso", elevation=4)
        layout.add_widget(toolbar)

        grid = MDGridLayout(cols=1, spacing=dp(20), padding=dp(40))

        buttons = [
            ("Paramètres", "cog", "settings"),
            ("Saisie Employés", "account-edit", "dataentry"),
            ("Payroll Data", "calculator", "payroll"),
            ("Journal Entries", "book-open", "journal"),
        ]

        for text, icon, screen in buttons:
            btn = MDRaisedButton(text=text, icon=icon, size_hint=(1, None), height=dp(60))
            btn.bind(on_release=lambda x, s=screen: self.manager.switch_screen(s))
            grid.add_widget(btn)

        layout.add_widget(grid)
        self.add_widget(layout)


class PaieScreenManager(MDScreenManager):
    def switch_screen(self, name):
        self.current = name


# ---------------------------------------------------------------------------
# APPLICATION PRINCIPALE
# ---------------------------------------------------------------------------
class PaieBurkinaApp(MDApp):
    settings = DEFAULT_SETTINGS.copy()
    employees = DEFAULT_EMPLOYEES.copy()


    def export_excel(self):
        """Exporte les données vers un fichier Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()

            # Settings
            ws_s = wb.active
            ws_s.title = "Settings"
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            params = [
                ["PARAMÈTRE", "VALEUR"],
                ["Mois", self.settings["mois"]],
                ["Année", self.settings["annee"]],
                ["CNSS Employé (%)", self.settings["cnss_employe_pct"]],
                ["Plafond CNSS", self.settings["plafond_cnss"]],
                ["Max CNSS Employé", self.settings["max_cnss_employe"]],
                ["Plafond Fiscal (%)", self.settings["plafond_fiscal_pct"]],
                ["Abattement Cadre (%)", self.settings["abattement_cadre"]],
                ["Abattement Non-Cadre (%)", self.settings["abattement_non_cadre"]],
                ["TPA (%)", self.settings["tpa_pct"]],
                ["CNSS Patronale (%)", self.settings["cnss_patronale_pct"]],
            ]
            for r_idx, row in enumerate(params, 1):
                for c_idx, val in enumerate(row, 1):
                    cell = ws_s.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font

            # Data Entry
            ws_d = wb.create_sheet("Data Entry")
            headers = ["N°", "Nom", "Prénom", "Classification", "Sal Base", "Prim anc", "Heure Sup",
                       "Sursalaire", "Gratif", "Indem", "Caisse", "Logement", "Fonction", "Transport",
                       "Charges", "Retenue Prêt"]
            for c_idx, h in enumerate(headers, 1):
                cell = ws_d.cell(row=1, column=c_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
            for r_idx, emp in enumerate(self.employees, 2):
                ws_d.cell(row=r_idx, column=1, value=emp["no"])
                ws_d.cell(row=r_idx, column=2, value=emp["nom"])
                ws_d.cell(row=r_idx, column=3, value=emp["prenom"])
                ws_d.cell(row=r_idx, column=4, value=emp["classification"])
                ws_d.cell(row=r_idx, column=5, value=emp["sal_base"])
                ws_d.cell(row=r_idx, column=6, value=emp["prim_anc"])
                ws_d.cell(row=r_idx, column=7, value=emp["heure_sup"])
                ws_d.cell(row=r_idx, column=8, value=emp["sursalaire"])
                ws_d.cell(row=r_idx, column=9, value=emp["gratif"])
                ws_d.cell(row=r_idx, column=10, value=emp["indem"])
                ws_d.cell(row=r_idx, column=11, value=emp["caisse"])
                ws_d.cell(row=r_idx, column=12, value=emp["logement"])
                ws_d.cell(row=r_idx, column=13, value=emp["fonction"])
                ws_d.cell(row=r_idx, column=14, value=emp["transport"])
                ws_d.cell(row=r_idx, column=15, value=emp["charges"])
                ws_d.cell(row=r_idx, column=16, value=emp["retenue_pret"])

            # Payroll Data
            ws_p = wb.create_sheet("Payroll Data")
            p_headers = ["N°", "Nom", "Prénom", "Rém Total", "CNSS", "Plafond Fiscal", "Sal Brut",
                         "Abattement", "Exonérations", "Base Imp", "IUTS Net", "Sal Net",
                         "Retenue 1%", "Retenue Prêt", "Net Perçu", "TPA", "CNSS Patronale",
                         "Total Charges", "Coûts"]
            for c_idx, h in enumerate(p_headers, 1):
                cell = ws_p.cell(row=1, column=c_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font

            results = [compute_employee(emp, self.settings) for emp in self.employees]
            for r_idx, r in enumerate(results, 2):
                ws_p.cell(row=r_idx, column=1, value=r["no"])
                ws_p.cell(row=r_idx, column=2, value=r["nom"])
                ws_p.cell(row=r_idx, column=3, value=r["prenom"])
                ws_p.cell(row=r_idx, column=4, value=r["rem_total"])
                ws_p.cell(row=r_idx, column=5, value=r["cnss_employe"])
                ws_p.cell(row=r_idx, column=6, value=r["plafond_fiscal"])
                ws_p.cell(row=r_idx, column=7, value=r["sal_brut"])
                ws_p.cell(row=r_idx, column=8, value=r["abattement"])
                ws_p.cell(row=r_idx, column=9, value=r["total_exon"])
                ws_p.cell(row=r_idx, column=10, value=r["base_imp"])
                ws_p.cell(row=r_idx, column=11, value=r["iuts_net"])
                ws_p.cell(row=r_idx, column=12, value=r["sal_net"])
                ws_p.cell(row=r_idx, column=13, value=r["retenue_1pct"])
                ws_p.cell(row=r_idx, column=14, value=r["retenue_pret"])
                ws_p.cell(row=r_idx, column=15, value=r["net_percu"])
                ws_p.cell(row=r_idx, column=16, value=r["tpa"])
                ws_p.cell(row=r_idx, column=17, value=r["cnss_patronale"])
                ws_p.cell(row=r_idx, column=18, value=r["total_charges_patronales"])
                ws_p.cell(row=r_idx, column=19, value=r["couts"])

            # Totaux
            totals = compute_totals(results)
            t_row = len(results) + 2
            ws_p.cell(row=t_row, column=1, value="TOTAL")
            for c_idx, key in enumerate(["rem_total", "cnss_employe", "plafond_fiscal", "sal_brut",
                                         "abattement", "total_exon", "base_imp", "iuts_net", "sal_net",
                                         "retenue_1pct", "retenue_pret", "net_percu", "tpa",
                                         "cnss_patronale", "total_charges_patronales", "couts"], 4):
                ws_p.cell(row=t_row, column=c_idx, value=totals[key])

            # Journal
            ws_j = wb.create_sheet("Journal Entries")
            journal = generate_journal_entries(totals)
            j_headers = ["Débit", "Crédit", "Libellé", "Montant"]
            for c_idx, h in enumerate(j_headers, 1):
                cell = ws_j.cell(row=1, column=c_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font

            r_idx = 2
            for e in journal["entries"]:
                ws_j.cell(row=r_idx, column=1, value=e["debit"])
                ws_j.cell(row=r_idx, column=2, value=e["credit"])
                ws_j.cell(row=r_idx, column=3, value=e["libelle"])
                ws_j.cell(row=r_idx, column=4, value=e["montant"])
                r_idx += 1

            path = os.path.join(self.get_data_dir(), f"Paie_{self.settings['mois']}_{self.settings['annee']}.xlsx")
            wb.save(path)
            return path
        except Exception as e:
            raise e

    def build(self):
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Light"
        self.load_data()

        sm = PaieScreenManager()
        sm.add_widget(MenuScreen(name="menu"))
        sm.add_widget(SettingsScreen(name="settings"))
        sm.add_widget(DataEntryScreen(name="dataentry"))
        sm.add_widget(PayrollScreen(name="payroll"))
        sm.add_widget(JournalScreen(name="journal"))
        return sm

    def get_data_dir(self):
        """Retourne le répertoire de données persistantes."""
        from kivy.storage.jsonstore import JsonStore
        # Sur Android : /sdcard/PaieBurkina
        # Sur desktop : ~/.paieburkina
        path = os.path.join(os.path.expanduser("~"), ".paieburkina")
        if not os.path.exists(path):
            os.makedirs(path)
        return path

    def save_data(self):
        path = self.get_data_dir()
        with open(os.path.join(path, "settings.json"), "w", encoding="utf-8") as f:
            json.dump(self.settings, f, indent=2, ensure_ascii=False)
        with open(os.path.join(path, "employees.json"), "w", encoding="utf-8") as f:
            json.dump(self.employees, f, indent=2, ensure_ascii=False)

    def load_data(self):
        path = self.get_data_dir()
        try:
            with open(os.path.join(path, "settings.json"), "r", encoding="utf-8") as f:
                self.settings = json.load(f)
        except:
            pass
        try:
            with open(os.path.join(path, "employees.json"), "r", encoding="utf-8") as f:
                self.employees = json.load(f)
        except:
            pass


if __name__ == "__main__":
    PaieBurkinaApp().run()
